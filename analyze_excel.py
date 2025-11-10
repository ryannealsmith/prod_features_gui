import pandas as pd
import openpyxl

# Load the Excel file
excel_file = 'Product Engineering Canonical Product Features.xlsx'
wb = openpyxl.load_workbook(excel_file)

print("Sheet names:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

print("\n" + "="*80 + "\n")

# Analyze each sheet
for sheet_name in wb.sheetnames:
    print(f"\nSheet: {sheet_name}")
    print("-" * 80)
    
    # Read with pandas for easier analysis
    df = pd.read_excel(excel_file, sheet_name=sheet_name)
    
    print(f"Shape: {df.shape[0]} rows x {df.shape[1]} columns")
    print(f"\nColumns:")
    for col in df.columns:
        print(f"  - {col}")
    
    print(f"\nFirst few rows:")
    print(df.head(3))
    print("\n" + "="*80)
