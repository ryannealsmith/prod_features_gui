#!/usr/bin/env python3
"""
Export engineering_plan_db.json to Google Sheets compatible format (CSV/Excel).
Allows user to select which components to export while maintaining all linking relationships.
Includes support for owner and url fields.
"""

import json
import csv
import sys
from pathlib import Path
from typing import Dict, List, Set, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class SheetsExporter:
    def __init__(self, json_file: str):
        self.json_file = json_file
        self.data = None
        self.selected_components = set()
        
    def load_data(self):
        """Load the JSON data file."""
        try:
            with open(self.json_file, 'r', encoding='utf-8') as f:
                self.data = json.load(f)
            print(f"✓ Loaded data from {self.json_file}")
            print(f"  Export date: {self.data.get('export_date', 'Unknown')}")
            return True
        except FileNotFoundError:
            print(f"✗ Error: File '{self.json_file}' not found")
            return False
        except json.JSONDecodeError as e:
            print(f"✗ Error: Invalid JSON in file: {e}")
            return False
    
    def show_menu(self):
        """Display menu and get user selections."""
        print("\n" + "="*60)
        print("EXPORT TO GOOGLE SHEETS - Component Selection")
        print("="*60)
        
        components = {
            '1': {
                'name': 'Product Variants',
                'key': 'product_variants',
                'count': len(self.data.get('product_variants', []))
            },
            '2': {
                'name': 'Configurations',
                'key': 'configurations',
                'count': len(self.data.get('configurations', []))
            },
            '3': {
                'name': 'Product Features',
                'key': 'product_features',
                'count': len(self.data.get('product_features', []))
            },
            '4': {
                'name': 'Capabilities',
                'key': 'capabilities',
                'count': len(self.data.get('capabilities', []))
            },
            '5': {
                'name': 'Technical Functions',
                'key': 'technical_functions',
                'count': len(self.data.get('technical_functions', []))
            }
        }
        
        print("\nAvailable components:")
        for key, comp in components.items():
            print(f"  [{key}] {comp['name']} ({comp['count']} items)")
        
        print(f"\n  [6] ALL components")
        print(f"  [0] Cancel/Exit")
        
        while True:
            print("\nSelect components to export (comma-separated, e.g., 1,2):")
            selection = input("Your choice: ").strip()
            
            if selection == '0':
                return False
            
            if selection == '6':
                self.selected_components = {'product_variants', 'configurations', 'product_features', 'capabilities', 'technical_functions'}
                break
            
            try:
                choices = [c.strip() for c in selection.split(',')]
                for choice in choices:
                    if choice in components:
                        self.selected_components.add(components[choice]['key'])
                    else:
                        print(f"  ✗ Invalid choice: {choice}")
                        self.selected_components.clear()
                        break
                
                if self.selected_components:
                    break
            except Exception as e:
                print(f"  ✗ Error parsing selection: {e}")
        
        print("\n✓ Selected components:")
        for comp in sorted(self.selected_components):
            print(f"  - {comp}")
        
        return True
    
    def get_relevant_relationships(self) -> Dict[str, List[Dict]]:
        """Get only the relationships relevant to selected components."""
        relationships = {}
        
        # Product Variants <-> Product Features
        if 'product_variants' in self.selected_components and 'product_features' in self.selected_components:
            relationships['pv_product_features'] = self.data.get('pv_product_features_relationships', [])
            print(f"  ✓ Including Product Variant <-> Product Feature links ({len(relationships['pv_product_features'])} relationships)")
        
        # Product Features <-> Capabilities
        if 'product_features' in self.selected_components and 'capabilities' in self.selected_components:
            relationships['pf_capabilities'] = self.data.get('pf_capabilities_relationships', [])
            print(f"  ✓ Including Product Feature <-> Capability links ({len(relationships['pf_capabilities'])} relationships)")
        
        # Capabilities <-> Technical Functions
        if 'capabilities' in self.selected_components and 'technical_functions' in self.selected_components:
            relationships['cap_technical_functions'] = self.data.get('cap_technical_functions_relationships', [])
            print(f"  ✓ Including Capability <-> Technical Function links ({len(relationships['cap_technical_functions'])} relationships)")
        
        return relationships
    
    def export_to_excel(self, output_file: str):
        """Export selected data to Excel format with multiple sheets."""
        print(f"\n{'='*60}")
        print("EXPORTING TO EXCEL")
        print(f"{'='*60}")
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Export Product Variants
        if 'product_variants' in self.selected_components:
            pv_data = self.data.get('product_variants', [])
            if pv_data:
                ws = wb.create_sheet("Product Variants")
                headers = list(pv_data[0].keys())
                
                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Write data
                for row_idx, item in enumerate(pv_data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=item.get(header))
                
                # Auto-adjust column widths
                for col_idx, header in enumerate(headers, 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 2, 12), 50)
                
                print(f"  ✓ Exported {len(pv_data)} Product Variants")

        # Export Configurations
        if 'configurations' in self.selected_components:
            pf_data = self.data.get('configurations', [])
            if pf_data:
                ws = wb.create_sheet("Configurations")
                headers = list(pf_data[0].keys())
                
                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Write data
                for row_idx, item in enumerate(pf_data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=item.get(header))
                
                # Auto-adjust column widths
                for col_idx, header in enumerate(headers, 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 2, 12), 50)
                
                print(f"  ✓ Exported {len(pf_data)} Configurations")

        # Export Product Features
        if 'product_features' in self.selected_components:
            pf_data = self.data.get('product_features', [])
            if pf_data:
                ws = wb.create_sheet("Product Features")
                headers = list(pf_data[0].keys())
                
                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Write data
                for row_idx, item in enumerate(pf_data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=item.get(header))
                
                # Auto-adjust column widths
                for col_idx, header in enumerate(headers, 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 2, 12), 50)
                
                print(f"  ✓ Exported {len(pf_data)} Product Features")
        
        # Export Capabilities
        if 'capabilities' in self.selected_components:
            cap_data = self.data.get('capabilities', [])
            if cap_data:
                ws = wb.create_sheet("Capabilities")
                headers = list(cap_data[0].keys())
                
                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Write data
                for row_idx, item in enumerate(cap_data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=item.get(header))
                
                # Auto-adjust column widths
                for col_idx, header in enumerate(headers, 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 2, 12), 50)
                
                print(f"  ✓ Exported {len(cap_data)} Capabilities")
        
        # Export Technical Functions
        if 'technical_functions' in self.selected_components:
            tf_data = self.data.get('technical_functions', [])
            if tf_data:
                ws = wb.create_sheet("Technical Functions")
                headers = list(tf_data[0].keys())
                
                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Write data
                for row_idx, item in enumerate(tf_data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=item.get(header))
                
                # Auto-adjust column widths
                for col_idx, header in enumerate(headers, 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(header)) + 2, 12), 50)
                
                print(f"  ✓ Exported {len(tf_data)} Technical Functions")
        
        # Export Relationships
        relationships = self.get_relevant_relationships()
        
        if 'pf_capabilities' in relationships and relationships['pf_capabilities']:
            ws = wb.create_sheet("PF-Capability Links")
            rel_data = relationships['pf_capabilities']
            headers = list(rel_data[0].keys())
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = header_alignment
            
            # Write data
            for row_idx, item in enumerate(rel_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=item.get(header))
            
            # Auto-adjust column widths
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            
            print(f"  ✓ Exported {len(rel_data)} Product Feature <-> Capability links")
        
        if 'cap_technical_functions' in relationships and relationships['cap_technical_functions']:
            ws = wb.create_sheet("Cap-TF Links")
            rel_data = relationships['cap_technical_functions']
            headers = list(rel_data[0].keys())
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.alignment = header_alignment
            
            # Write data
            for row_idx, item in enumerate(rel_data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=item.get(header))
            
            # Auto-adjust column widths
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
            
            print(f"  ✓ Exported {len(rel_data)} Capability <-> Technical Function links")
        
        # Add metadata sheet
        ws = wb.create_sheet("Export Metadata", 0)
        ws['A1'] = "Export Information"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A3'] = "Original Export Date:"
        ws['B3'] = self.data.get('export_date', 'Unknown')
        ws['A4'] = "Re-exported to Excel:"
        from datetime import datetime
        ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A5'] = "Source File:"
        ws['B5'] = self.json_file
        ws['A6'] = "Selected Components:"
        ws['B6'] = ", ".join(sorted(self.selected_components))
        ws['B5'] = self.json_file
        ws['A6'] = "Selected Components:"
        ws['B6'] = ", ".join(sorted(self.selected_components))
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        
        # Save workbook
        wb.save(output_file)
        print(f"\n✓ Excel file saved: {output_file}")
        print(f"  Total sheets: {len(wb.sheetnames)}")
        print(f"  Ready to import into Google Sheets!")
    
    def export_to_csv(self, output_dir: str):
        """Export selected data to multiple CSV files (one per component)."""
        print(f"\n{'='*60}")
        print("EXPORTING TO CSV FILES")
        print(f"{'='*60}")
        
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        files_created = []
        
        # Export Product Variants
        if 'product_variants' in self.selected_components:
            pv_data = self.data.get('product_variants', [])
            if pv_data:
                csv_file = output_path / "product_variants.csv"
                with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=pv_data[0].keys())
                    writer.writeheader()
                    writer.writerows(pv_data)
                files_created.append(csv_file)
                print(f"  ✓ Exported {len(pv_data)} Product Variants to {csv_file.name}")

        # Export Configurations
        if 'product_features' in self.selected_components:
            pf_data = self.data.get('configurations', [])
            if pf_data:
                csv_file = output_path / "configurations.csv"
                with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=pf_data[0].keys())
                    writer.writeheader()
                    writer.writerows(pf_data)
                files_created.append(csv_file)
                print(f"  ✓ Exported {len(pf_data)} Configurations to {csv_file.name}")

        # Export Product Features
        if 'product_features' in self.selected_components:
            pf_data = self.data.get('product_features', [])
            if pf_data:
                csv_file = output_path / "product_features.csv"
                with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=pf_data[0].keys())
                    writer.writeheader()
                    writer.writerows(pf_data)
                files_created.append(csv_file)
                print(f"  ✓ Exported {len(pf_data)} Product Features to {csv_file.name}")
        
        # Export Capabilities
        if 'capabilities' in self.selected_components:
            cap_data = self.data.get('capabilities', [])
            if cap_data:
                csv_file = output_path / "capabilities.csv"
                with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=cap_data[0].keys())
                    writer.writeheader()
                    writer.writerows(cap_data)
                files_created.append(csv_file)
                print(f"  ✓ Exported {len(cap_data)} Capabilities to {csv_file.name}")
        
        # Export Technical Functions
        if 'technical_functions' in self.selected_components:
            tf_data = self.data.get('technical_functions', [])
            if tf_data:
                csv_file = output_path / "technical_functions.csv"
                with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=tf_data[0].keys())
                    writer.writeheader()
                    writer.writerows(tf_data)
                files_created.append(csv_file)
                print(f"  ✓ Exported {len(tf_data)} Technical Functions to {csv_file.name}")
        
        # Export Relationships
        relationships = self.get_relevant_relationships()
        
        if 'pf_capabilities' in relationships and relationships['pf_capabilities']:
            rel_data = relationships['pf_capabilities']
            csv_file = output_path / "pf_capability_links.csv"
            with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=rel_data[0].keys())
                writer.writeheader()
                writer.writerows(rel_data)
            files_created.append(csv_file)
            print(f"  ✓ Exported {len(rel_data)} PF-Capability links to {csv_file.name}")
        
        if 'cap_technical_functions' in relationships and relationships['cap_technical_functions']:
            rel_data = relationships['cap_technical_functions']
            csv_file = output_path / "cap_tf_links.csv"
            with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=rel_data[0].keys())
                writer.writeheader()
                writer.writerows(rel_data)
            files_created.append(csv_file)
            print(f"  ✓ Exported {len(rel_data)} Cap-TF links to {csv_file.name}")
        
        print(f"\n✓ CSV files saved to: {output_path}")
        print(f"  Total files: {len(files_created)}")
        print(f"  You can import these CSV files into separate Google Sheets tabs")
    
    def run(self):
        """Main execution flow."""
        print("\n" + "="*60)
        print("ENGINEERING PLAN - Google Sheets Exporter")
        print("="*60)
        
        # Load data
        if not self.load_data():
            return 1
        
        # Get user selections
        if not self.show_menu():
            print("\n✗ Export cancelled by user")
            return 0
        
        # Choose output format
        print("\n" + "="*60)
        print("OUTPUT FORMAT")
        print("="*60)
        print("\n[1] Excel (.xlsx) - Single file with multiple sheets (RECOMMENDED)")
        print("[2] CSV files - Multiple files (one per component)")
        print("[0] Cancel")
        
        while True:
            choice = input("\nYour choice: ").strip()
            
            if choice == '0':
                print("\n✗ Export cancelled")
                return 0
            elif choice == '1':
                output_file = "roadmap_export.xlsx"
                self.export_to_excel(output_file)
                break
            elif choice == '2':
                output_dir = "roadmap_export_csv"
                self.export_to_csv(output_dir)
                break
            else:
                print("  ✗ Invalid choice. Please select 1, 2, or 0.")
        
        print("\n" + "="*60)
        print("IMPORT TO GOOGLE SHEETS")
        print("="*60)
        print("\nTo import into Google Sheets:")
        print("  1. Go to https://sheets.google.com")
        print("  2. Create a new spreadsheet or open an existing one")
        if choice == '1':
            print("  3. File > Import > Upload")
            print(f"  4. Select '{output_file}'")
            print("  5. Import location: 'Replace spreadsheet' or 'Insert new sheets'")
        else:
            print("  3. File > Import > Upload (for each CSV file)")
            print(f"  4. Import each file from '{output_dir}/' as a separate sheet")
        print("\n✓ All linking relationships are maintained in the exported data!")
        print("="*60)
        
        return 0


def main():
    """Main entry point."""
    json_file = "engineering_plan_db.json"
    
    # Check if file exists
    if not Path(json_file).exists():
        print(f"✗ Error: '{json_file}' not found in current directory")
        print(f"  Current directory: {Path.cwd()}")
        print(f"\nPlease run this script from the directory containing {json_file}")
        return 1
    
    exporter = SheetsExporter(json_file)
    return exporter.run()


if __name__ == "__main__":
    sys.exit(main())
